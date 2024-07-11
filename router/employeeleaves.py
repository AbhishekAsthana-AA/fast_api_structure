import html
from pipes import Template
import re
from fastapi import Depends, HTTPException, APIRouter
from sqlalchemy.orm import Session
from sqlalchemy.sql import text
from openpyxl import Workbook
from typing import List
from database.db import get_db
# from models.employee_modal import employeeLeaveDetails, employeeLoginDetails
from schema.employee_details import employeeLeaves,employeeLoginPayload,deleteLoginPayload,batchInsert
from crud.employeeleavescrud import getEmployeeLeaves,insertEmpDetails,deleteEmpDetails,getEmpdata,getDatabyjoin, batchresult
from datetime import datetime
from starlette.responses import FileResponse
import time
import pdfkit
from jinja2 import Environment, FileSystemLoader,Template
import os
from fastapi.responses import JSONResponse
from fastapi.encoders import jsonable_encoder



router = APIRouter()

@router.post('/getempleaves')
def getEmpLeaves(emp_details: employeeLeaves, db: Session = Depends(get_db)):
    try:
        result = getEmployeeLeaves(db, emp_details)
        if not result: 
            return JSONResponse(status_code=400, content={"detail": "Data not found"})
        return JSONResponse(status_code=200, content={"detail": "Data found", "value":jsonable_encoder(result)}) 
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LINE: {e.__traceback__.tb_lineno} ERROR: {str(e)}")

@router.post('/generateExcel')
async def generateExcel(emp_details : employeeLeaves , db:Session = Depends(get_db)):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Sheet"
    headers = ["Total Days", "Start Date", "End Date", "ID", "Employee ID", "Leave ID"]
    result = getEmployeeLeaves(db, emp_details)

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row_data in enumerate(result, start=2):
        ws.cell(row=row_idx, column=1).value = row_data.employee_id

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_file = os.path.join(base_dir, 'src', 'uploads','downloads')
    file_path = os.path.join(excel_file, 'sample.xlsx')
    wb.save(file_path)
    res= FileResponse(path=file_path, filename="sample.xlsx")
    def remove_file():
        os.remove(file_path)
    return res


@router.post('/readpdf')
async def readtemplate(emp_details : employeeLeaves , db:Session = Depends(get_db)):
    # C:\Program Files\wkhtmltopdf
    try:
        path_to_wkhtmltopdf = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe' 
        config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf)
        result = getEmployeeLeaves(db, emp_details)
        load_temp= os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'leave.html')
        with open(load_temp) as file:
            template = Template(file.read())
        html_content = template.render(users=result)

        filename = f"pdfgenerate_{int(time.time())}.pdf"
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pdf_dir = os.path.join(base_dir, 'src', 'uploads','downloads')
        os.makedirs(pdf_dir, exist_ok=True)
        file_path = os.path.join(pdf_dir, filename)
        pdfkit.from_string(html_content, file_path,configuration=config)
        if not os.path.exists(file_path):
          raise HTTPException(status_code=500, detail="Failed to generate PDF file")
        res=FileResponse(path=file_path, filename=filename)
        return  res 
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LINE: {e.__traceback__.tb_lineno} ERROR: {str(e)}")

@router.post('/createempdetails')
def insertUpdateEmpLeaves(employee_data: employeeLoginPayload, db:Session = Depends(get_db)):
    try:
        result = insertEmpDetails(db,employee_data)
        return  result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LINE: {e.__traceback__.tb_lineno} ERROR: {str(e)}")

@router.post('/delete')
def deleteEmpLeaves(employee_id: deleteLoginPayload, db:Session = Depends(get_db)):
    try:
        result = deleteEmpDetails(db,employee_id)
        return  result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LINE: {e.__traceback__.tb_lineno} ERROR: {str(e)}")
    
@router.post('/getempdata')
def getEmpLeaves(employee_id: deleteLoginPayload, db:Session = Depends(get_db)):
    try:
        result = getEmpdata(db,employee_id)
        return JSONResponse(status_code = 200, content=result)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LINE: {e.__traceback__.tb_lineno} ERROR: {str(e)}")
    
@router.post('/getdatawithjoin')
def getJoinData(employee_id: deleteLoginPayload, db:Session = Depends(get_db)):
    try:
        result = getDatabyjoin(db, employee_id)
        if result is None:
            return JSONResponse(status_code=200, content={"result":"No data"}) 
        else:
            return JSONResponse(status_code=200, content={"message":"Data Found","status":"TRUE","value":jsonable_encoder(result)})
           
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LINE: {e.__traceback__.tb_lineno} ERROR: {str(e)}")
    

@router.post('/getbatchinsert')
def getatchData(batchData: batchInsert, db:Session = Depends(get_db)):
    try:
        result = batchresult(db, batchData.data)
        if result is None:
            return JSONResponse(status_code=200, content={"result":"No data"}) 
        else:
            return JSONResponse(status_code=200, content={"message":"Data Found","status":"TRUE","value":jsonable_encoder(result)})
           
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LINE: {e.__traceback__.tb_lineno} ERROR: {str(e)}")
 
 